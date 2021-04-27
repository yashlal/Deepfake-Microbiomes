import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import pickle
import seaborn as sb
from predict_by_model import *
import openpyxl
import random as rd

def genrand(p):
    v = rd.random()
    if v <= p:
        return 1
    else:
        return 0

def GetStrn(strg):
    s1 = strg.split(";")[-1]
    s2 = "_".join(s1.split("__")[1:])
    return s2

def get_all_species(csv, _sep):

    miceData = pd.read_csv(csv, sep=_sep)
    species = miceData.species.apply(GetStrn)
    species = list(species)
    species = species[:-1]
    #remove blank element
    return species

#generate a random excel for testing
def edit_pairwise_data(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb['Relative_Abundance']

    for i in range(2, sheet.max_row+1):
        for j in range(2, sheet.max_column+1):
            print(i)
            if i==j:
                sheet.cell(row=i,column=j).value = 1
            else:
                sheet.cell(row=i,column=j).value = rd.random()
    wb.save('rand.xlsx')

def make_prob_distro(species_list):
    prob_dict = {}
    for i in range(len(species_list)):
        prob_dict[species_list[i]] = 1
    return prob_dict

def generator(prob_dictionary, pairwise_file, n):
    CommunityEquilibrium = {}
    for i in range(0,n):
        print("Trial " + str(i+1) + " in progress...")
        spec_list = []

        for species in prob_dictionary:
            bin = genrand(prob_dictionary[species])
            if bin == 1:
                spec_list.append(species)

        Equilibrium, FoundList = predict_community(spec_list, File = pairwise_file, lambdaVersion = "Equilibrium", verb = True)

        CommunityEquilibrium[i+1] = dict([(ky,val.round(3)) for ky,val in Equilibrium.items()])
    return CommunityEquilibrium, FoundList

def update_pw(file, unfound):
    wb = openpyxl.load_workbook(file)
    sh = wb['Relative_Abundance']
    mr, mc = sh.max_row, sh.max_column
    for i in range(len(unfound)):
        sh.cell(row=mr+i+1, column=1).value = unfound[i]
        sh.cell(row=1, column = mc+i+1).value = unfound[i]
    wb.save('updated.xlsx')

s = get_all_species(csv="Cdiff_mice_high_vs_low_risk.species.tsv", _sep='\t')
true_prob_distro = make_prob_distro(s)
CU, FL = generator(true_prob_distro, 'updated.xlsx', 1)
print(CU)
