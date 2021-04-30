import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import pickle
import seaborn as sb
from predict_by_model import *
import openpyxl
import random as rd

def fileopener(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    sh = wb[sheet_name]
    return wb, sh

def gen_random_matrix(wb, sheet_object):
    i = 2
    while i <= sheet_object.max_row:
        print('On row ' + str(i) + ' out of ' + str(sheet_object.max_row))
        for j in range(2,i+1):
            if i == j:
                sheet_object.cell(row=i, column=j).value = 1
            else:
                x = rd.random()
                sheet_object.cell(row=i,column=j).value = x
                sheet_object.cell(row=j, column = i).value = 1-x
        i += 1
    wb.save('rand.xlsx')
