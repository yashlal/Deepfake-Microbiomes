import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
import pickle
import seaborn as sb
from predict_by_model import *
import openpyxl
import random as rd

def genrand(p):
    v = rd.random()
    if v <= p:
        return 1
    else:
        return 0

def GetStrn(strg):
    s1 = strg.split(";")[-1]
    s2 = "_".join(s1.split("__")[1:])
    return s2

def model_comm_eq(csv, _sep, dirhead, ObservationFile):

    miceData = pd.read_csv(csv, sep=_sep)
    species = miceData.species.apply(GetStrn)
    species = list(species)
    species = species[:-1]
    #remove blank element
    return species

#generate a random excel for testing
def edit_pairwise_data(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb['Relative_Abundance']

    for i in range(2, sheet.max_row+1):
        for j in range(2, sheet.max_column+1):
            if i==j:
                sheet.cell(row=i,column=j).value = 1
            else:
                sheet.cell(row=i,column=j).value = rd.random()
    wb.save('rand.xlsx')

s = model_comm_eq(csv="Cdiff_mice_high_vs_low_risk.species.tsv", _sep='\t', dirhead='T', ObservationFile =  "rand.xlsx")
print(s)
prob_dict = {}

#makes a random prob distro
for i in range(len(s)):
    prob_dict[s[i]] = rd.random()

print(prob_dict)

def generator(prob_dictionary, pairwise_file, n):
    CommunityEquilibrium = {}
    for i in range(0,n):
        spec_list = []

        for species in prob_dictionary:
            bin = genrand(prob_dictionary[species])
            if bin == 1:
                spec_list.append(species)

        Equilibrium, FoundList = predict_community(spec_list, File = pairwise_file, lambdaVersion = "Equilibrium", verb = True)

        CommunityEquilibrium = dict([(ky,val.round(3)) for ky,val in Equilibrium.items()])

        d = pd.DataFrame(list(CommunityEquilibrium.items()))
        d.to_excel("new.xlsx")

generator(prob_dict, 'Pairwise_Chemostat.xlsx', 1)
